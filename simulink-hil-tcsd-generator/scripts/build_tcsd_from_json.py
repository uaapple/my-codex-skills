#!/usr/bin/env python3
"""Build a TCSD workbook from a compact JSON spec and the bundled template."""

from __future__ import annotations

import argparse
import json
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


HEADERS = [
    "TestID",
    "Name",
    "Type",
    "Requirement ID",
    "Test Case Description",
    "Initialization",
    "Action",
    "Work Status",
    "Report Links",
]


def copy_row_style(ws, src_row: int, dst_row: int) -> None:
    for col in range(1, 10):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--template", required=True)
    parser.add_argument("--spec", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    spec = json.loads(Path(args.spec).read_text(encoding="utf-8"))
    wb = load_workbook(args.template)
    ws = wb["TCSD"]

    for col, header in enumerate(HEADERS, 1):
        ws.cell(1, col).value = header

    group = spec.get("test_group", {})
    ws.cell(2, 1).value = group.get("id", "INFO_TG_001")
    ws.cell(2, 2).value = group.get("name", spec.get("model_name", "Model"))
    ws.cell(2, 3).value = "TestGroup"
    ws.cell(2, 4).value = "TestGroup"
    ws.cell(2, 5).value = group.get("description", "")
    ws.cell(2, 6).value = group.get("initialization_1", "")
    ws.cell(3, 6).value = group.get("initialization_2", "")
    ws.cell(4, 6).value = group.get("output_reference_1", "")
    ws.cell(5, 6).value = group.get("output_reference_2", "")

    for row in range(6, max(ws.max_row, 40) + 1):
        for col in range(1, 10):
            ws.cell(row, col).value = None

    for row, test in enumerate(spec["tests"], start=6):
        copy_row_style(ws, 6 + ((row - 6) % 5), row)
        ws.cell(row, 1).value = test["id"]
        ws.cell(row, 2).value = test["name"]
        ws.cell(row, 3).value = "Test"
        ws.cell(row, 4).value = test.get("requirement_id")
        ws.cell(row, 5).value = test.get("description", "")
        ws.cell(row, 6).value = test.get("initialization", "")
        ws.cell(row, 7).value = test.get("action", "")
        ws.cell(row, 8).value = test.get("work_status", "reviewed")
        ws.cell(row, 9).value = test.get("report_links")
        line_count = (ws.cell(row, 7).value or "").count("\n") + 1
        ws.row_dimensions[row].height = min(409, max(180, line_count * 13))

    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
