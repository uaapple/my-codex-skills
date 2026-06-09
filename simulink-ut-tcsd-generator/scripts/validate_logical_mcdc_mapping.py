#!/usr/bin/env python3
"""Validate that a TCSD workbook contains required Logical Operator MC/DC states.

This is a workbook-mapping gate. It proves that TCSD root-input/parameter
assignments contain the model-derived MC/DC vectors. It does not replace a
Simulink Coverage report or internal block probe, but it prevents a workbook
from silently omitting required AND/OR vectors.
"""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
from copy import deepcopy
from pathlib import Path
from typing import Any

STEP_RE = re.compile(r"^\s*\[\+\s*([0-9.]+)\s*(ms|s)\s*\].*$", re.IGNORECASE)
NUMBER_RE = r"[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?"
VALUE_RE = rf"({NUMBER_RE}|true|false|TRUE|FALSE|[A-Za-z_]\w*)"
PARAM_ASSIGN_RE = re.compile(rf"^\s*p\s+([A-Za-z_]\w*)\s*=\s*{VALUE_RE}\s*;")
ASSIGN_RE = re.compile(rf"^\s*([A-Za-z_]\w*)\s*=\s*{VALUE_RE}\s*;")
INDEX_ASSIGN_RE = re.compile(rf"^\s*([A-Za-z_]\w*)\s+([1-9]\d*)\s*=\s*{VALUE_RE}\s*;")
VECTOR_ASSIGN_RE = re.compile(r"^\s*([A-Za-z_]\w*)\s*=\s*\[([^\]]+)\]\s*;")


def header_index(ws, header: str, fallback: int) -> int:
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(1, col).value or "").strip() == header:
            return col
    return fallback


def parse_value(raw: Any) -> Any:
    if isinstance(raw, (int, float, bool)):
        return raw
    text = str(raw).strip()
    if text.lower() == "true":
        return True
    if text.lower() == "false":
        return False
    try:
        return float(text)
    except ValueError:
        return text


def normalize_key(raw_key: str) -> str:
    key = str(raw_key).strip()
    if not key:
        return key
    if "[" in key and key.endswith("]"):
        base, index = key[:-1].split("[", 1)
        return f"{base.strip()}[{index.strip()}]"
    parts = key.split()
    if len(parts) == 2 and parts[1].isdigit():
        return f"{parts[0]}[{parts[1]}]"
    return key


def normalize_expected_map(value_map: Any) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    if not isinstance(value_map, dict):
        return normalized
    for raw_key, raw_value in value_map.items():
        key = normalize_key(str(raw_key))
        if isinstance(raw_value, list):
            for index, item in enumerate(raw_value, start=1):
                normalized[f"{key}[{index}]"] = parse_value(item)
        elif isinstance(raw_value, dict) and raw_value and all(str(k).isdigit() for k in raw_value):
            for raw_index, item in raw_value.items():
                normalized[f"{key}[{raw_index}]"] = parse_value(item)
        else:
            normalized[key] = parse_value(raw_value)
    return normalized


def values_match(actual: Any, expected: Any, tolerance: float) -> bool:
    if isinstance(actual, bool) or isinstance(expected, bool):
        return actual is expected
    if isinstance(actual, (int, float)) and isinstance(expected, (int, float)):
        return math.isclose(float(actual), float(expected), rel_tol=0.0, abs_tol=tolerance)
    return actual == expected


def parse_assignment_line(line: str) -> tuple[str, str, Any] | None:
    stripped = line.strip()
    if not stripped or stripped.startswith("//"):
        return None
    param_match = PARAM_ASSIGN_RE.match(stripped)
    if param_match:
        return "params", param_match.group(1), parse_value(param_match.group(2))
    index_match = INDEX_ASSIGN_RE.match(stripped)
    if index_match:
        return "inputs", f"{index_match.group(1)}[{index_match.group(2)}]", parse_value(index_match.group(3))
    vector_match = VECTOR_ASSIGN_RE.match(stripped)
    if vector_match:
        # Return a sentinel; parse_assignments expands it.
        return "vector", vector_match.group(1), vector_match.group(2)
    assign_match = ASSIGN_RE.match(stripped)
    if assign_match:
        return "inputs", assign_match.group(1), parse_value(assign_match.group(2))
    return None


def parse_assignments(text: str) -> tuple[dict[str, Any], dict[str, Any]]:
    inputs: dict[str, Any] = {}
    params: dict[str, Any] = {}
    for raw in (text or "").replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        parsed = parse_assignment_line(raw)
        if not parsed:
            continue
        kind, key, value = parsed
        if kind == "params":
            params[key] = value
        elif kind == "vector":
            values = [item for item in re.split(r"[\s,]+", str(value).strip()) if item]
            for index, item in enumerate(values, start=1):
                inputs[f"{key}[{index}]"] = parse_value(item)
        else:
            inputs[key] = value
    return inputs, params


def merge_state(
    base_inputs: dict[str, Any],
    base_params: dict[str, Any],
    next_inputs: dict[str, Any],
    next_params: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    inputs = dict(base_inputs)
    params = dict(base_params)
    inputs.update(next_inputs)
    params.update(next_params)
    return inputs, params


def snapshot(
    *,
    test_id: str,
    row: int,
    name: str,
    phase: str,
    step_index: int | None,
    inputs: dict[str, Any],
    params: dict[str, Any],
) -> dict[str, Any]:
    return {
        "test_id": test_id,
        "row": row,
        "name": name,
        "phase": phase,
        "step_index": step_index,
        "inputs": deepcopy(inputs),
        "params": deepcopy(params),
    }


def parse_action_snapshots(
    action: str,
    *,
    test_id: str,
    row: int,
    name: str,
    init_inputs: dict[str, Any],
    init_params: dict[str, Any],
) -> list[dict[str, Any]]:
    snapshots: list[dict[str, Any]] = []
    current_inputs = dict(init_inputs)
    current_params = dict(init_params)
    step_index = 0
    pending_lines: list[str] = []

    def flush_step() -> None:
        nonlocal current_inputs, current_params, pending_lines
        if step_index == 0:
            pending_lines = []
            return
        step_inputs: dict[str, Any] = {}
        step_params: dict[str, Any] = {}
        for line in pending_lines:
            inputs, params = parse_assignments(line)
            step_inputs.update(inputs)
            step_params.update(params)
        current_inputs, current_params = merge_state(current_inputs, current_params, step_inputs, step_params)
        snapshots.append(
            snapshot(
                test_id=test_id,
                row=row,
                name=name,
                phase="action_step",
                step_index=step_index,
                inputs=current_inputs,
                params=current_params,
            )
        )
        pending_lines = []

    for raw in (action or "").splitlines():
        if STEP_RE.match(raw):
            flush_step()
            step_index += 1
            continue
        pending_lines.append(raw)
    flush_step()
    return snapshots


def workbook_snapshots(workbook: str | Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "openpyxl is required to read TCSD workbooks; use the bundled Codex Python runtime "
            "or install openpyxl in the active Python environment"
        ) from exc

    wb = load_workbook(workbook, data_only=False)
    if "TCSD" not in wb.sheetnames:
        raise ValueError("workbook does not contain a TCSD sheet")
    ws = wb["TCSD"]
    test_id_col = header_index(ws, "TestID", 1)
    name_col = header_index(ws, "Name", 2)
    type_col = header_index(ws, "Type", 3)
    init_col = header_index(ws, "Initialization", 6)
    action_col = header_index(ws, "Action", 7)

    group_inputs: dict[str, Any] = {}
    group_params: dict[str, Any] = {}
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, type_col).value or "").strip() != "TestGroup":
            continue
        first_inputs, first_params = parse_assignments(str(ws.cell(row, init_col).value or ""))
        second_inputs, second_params = parse_assignments(str(ws.cell(row + 1, init_col).value or ""))
        group_inputs, group_params = merge_state(first_inputs, first_params, second_inputs, second_params)
        break

    snapshots: list[dict[str, Any]] = []
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, type_col).value or "").strip() != "Test":
            continue
        test_id = str(ws.cell(row, test_id_col).value or "").strip()
        name = str(ws.cell(row, name_col).value or "").strip()
        test_inputs, test_params = parse_assignments(str(ws.cell(row, init_col).value or ""))
        init_inputs, init_params = merge_state(group_inputs, group_params, test_inputs, test_params)
        snapshots.append(
            snapshot(
                test_id=test_id,
                row=row,
                name=name,
                phase="initialization",
                step_index=None,
                inputs=init_inputs,
                params=init_params,
            )
        )
        snapshots.extend(
            parse_action_snapshots(
                str(ws.cell(row, action_col).value or ""),
                test_id=test_id,
                row=row,
                name=name,
                init_inputs=init_inputs,
                init_params=init_params,
            )
        )
    return snapshots


def obligation_status(obligation: dict[str, Any]) -> str:
    return str(obligation.get("status") or "required").strip().lower()


def obligation_matches(obligation: dict[str, Any]) -> list[dict[str, Any]]:
    if isinstance(obligation.get("matches"), list):
        return [item for item in obligation["matches"] if isinstance(item, dict)]
    if isinstance(obligation.get("match"), dict):
        return [obligation["match"]]
    return [
        {
            "inputs": obligation.get("expected_inputs")
            or obligation.get("expected_assignments")
            or obligation.get("inputs")
            or {},
            "params": obligation.get("expected_params")
            or obligation.get("expected_parameters")
            or obligation.get("params")
            or {},
        }
    ]


def normalize_match(match: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any]]:
    inputs = normalize_expected_map(
        match.get("inputs") or match.get("expected_inputs") or match.get("expected_assignments") or {}
    )
    params = normalize_expected_map(
        match.get("params") or match.get("parameters") or match.get("expected_params") or {}
    )
    # Allow compact "p Param" keys inside inputs/expected_assignments.
    for key in list(inputs):
        if key.startswith("p "):
            params[key[2:].strip()] = inputs.pop(key)
    return inputs, params


def missing_items(
    snap: dict[str, Any],
    expected_inputs: dict[str, Any],
    expected_params: dict[str, Any],
    tolerance: float,
) -> list[dict[str, Any]]:
    missing: list[dict[str, Any]] = []
    for key, expected in expected_inputs.items():
        if key not in snap["inputs"]:
            missing.append({"kind": "input", "key": key, "expected": expected, "actual": None})
        elif not values_match(snap["inputs"][key], expected, tolerance):
            missing.append({"kind": "input", "key": key, "expected": expected, "actual": snap["inputs"][key]})
    for key, expected in expected_params.items():
        if key not in snap["params"]:
            missing.append({"kind": "param", "key": key, "expected": expected, "actual": None})
        elif not values_match(snap["params"][key], expected, tolerance):
            missing.append({"kind": "param", "key": key, "expected": expected, "actual": snap["params"][key]})
    return missing


def find_match(
    obligation: dict[str, Any],
    snapshots: list[dict[str, Any]],
    *,
    tolerance: float,
    require_planned_test: bool,
) -> tuple[dict[str, Any] | None, list[dict[str, Any]]]:
    candidates = snapshots
    planned_test_id = obligation.get("planned_test_id") or obligation.get("test_id")
    if require_planned_test and planned_test_id:
        candidates = [snap for snap in snapshots if snap.get("test_id") == planned_test_id]

    closest_missing: list[dict[str, Any]] = []
    for match in obligation_matches(obligation):
        expected_inputs, expected_params = normalize_match(match)
        for snap in candidates:
            missing = missing_items(snap, expected_inputs, expected_params, tolerance)
            if not missing:
                return snap, []
            if not closest_missing or len(missing) < len(closest_missing):
                closest_missing = missing
    return None, closest_missing


def load_obligations(path: str | Path) -> list[dict[str, Any]]:
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    obligations = data.get("obligations") if isinstance(data, dict) else data
    if not isinstance(obligations, list):
        raise ValueError("obligations JSON must contain an obligations list")
    return [item for item in obligations if isinstance(item, dict)]


def validate_mapping(
    workbook: str | Path,
    obligations_path: str | Path,
    *,
    tolerance: float = 1e-9,
    require_planned_test: bool = False,
) -> dict[str, Any]:
    snapshots = workbook_snapshots(workbook)
    obligations = load_obligations(obligations_path)
    covered: list[dict[str, Any]] = []
    missing: list[dict[str, Any]] = []
    waived: list[dict[str, Any]] = []
    unresolved: list[dict[str, Any]] = []

    for obligation in obligations:
        status = obligation_status(obligation)
        item_id = obligation.get("id")
        if status in {"unreachable", "waived", "not_traceable"}:
            waived.append(
                {
                    "id": item_id,
                    "status": status,
                    "reason": obligation.get("reason") or obligation.get("unreachable_reason"),
                    "block_path": obligation.get("block_path"),
                }
            )
            continue
        if status in {"unresolved", "needs_manual_resolution"}:
            unresolved.append(
                {
                    "id": item_id,
                    "status": status,
                    "block_path": obligation.get("block_path"),
                    "issues": obligation.get("issues", []),
                }
            )
            continue
        matched, closest = find_match(
            obligation,
            snapshots,
            tolerance=tolerance,
            require_planned_test=require_planned_test,
        )
        if matched:
            covered.append(
                {
                    "id": item_id,
                    "block_path": obligation.get("block_path"),
                    "required_outcome": obligation.get("required_outcome"),
                    "matched": {
                        "test_id": matched.get("test_id"),
                        "row": matched.get("row"),
                        "phase": matched.get("phase"),
                        "step_index": matched.get("step_index"),
                    },
                }
            )
        else:
            missing.append(
                {
                    "id": item_id,
                    "block_path": obligation.get("block_path"),
                    "required_outcome": obligation.get("required_outcome"),
                    "closest_missing": closest,
                }
            )

    status = "passed" if not missing and not unresolved else "failed"
    return {
        "status": status,
        "workbook": str(workbook),
        "obligations": str(obligations_path),
        "summary": {
            "snapshot_count": len(snapshots),
            "obligation_count": len(obligations),
            "covered_count": len(covered),
            "missing_count": len(missing),
            "unresolved_count": len(unresolved),
            "waived_count": len(waived),
        },
        "covered": covered,
        "missing": missing,
        "unresolved": unresolved,
        "waived": waived,
    }


def print_text_report(report: dict[str, Any]) -> None:
    summary = report.get("summary", {})
    print(
        "Logical MC/DC mapping validation "
        f"{report.get('status')}: "
        f"{summary.get('covered_count', 0)} covered, "
        f"{summary.get('missing_count', 0)} missing, "
        f"{summary.get('unresolved_count', 0)} unresolved, "
        f"{summary.get('waived_count', 0)} waived."
    )
    for item in report.get("missing", [])[:30]:
        print(f"- missing {item.get('id')}: {item.get('block_path')}", file=sys.stderr)
        if item.get("required_outcome"):
            print(f"  {item['required_outcome']}", file=sys.stderr)
        for missing in item.get("closest_missing", [])[:8]:
            print(
                f"  {missing.get('kind')} {missing.get('key')}: "
                f"expected {missing.get('expected')}, actual {missing.get('actual')}",
                file=sys.stderr,
            )
    for item in report.get("unresolved", [])[:30]:
        print(f"- unresolved {item.get('id')}: {item.get('block_path')}", file=sys.stderr)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--obligations", required=True, help="coverage_obligations.json path")
    parser.add_argument("--report-json")
    parser.add_argument("--tolerance", type=float, default=1e-9)
    parser.add_argument(
        "--require-planned-test",
        action="store_true",
        help="If an obligation declares planned_test_id/test_id, only that Test can satisfy it.",
    )
    args = parser.parse_args()

    report = validate_mapping(
        args.workbook,
        args.obligations,
        tolerance=args.tolerance,
        require_planned_test=args.require_planned_test,
    )
    if args.report_json:
        Path(args.report_json).write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    else:
        print_text_report(report)
    return 0 if report["status"] == "passed" else 1


if __name__ == "__main__":
    raise SystemExit(main())
