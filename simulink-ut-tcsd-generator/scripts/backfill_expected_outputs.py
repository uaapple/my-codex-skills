#!/usr/bin/env python3
"""Backfill TCSD Action cells with simulation results for root outputs only."""

from __future__ import annotations

import argparse
import json
import re
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


STEP_RE = re.compile(r"^\s*\[\+")
ANY_EXPECTED_RE = re.compile(r"^\s*([A-Za-z_]\w*)\s*=\s*expValue\(")


def parse_steps(action: str) -> list[dict]:
    steps: list[dict] = []
    current: dict | None = None
    for raw in (action or "").splitlines():
        if STEP_RE.match(raw):
            if current is not None:
                steps.append(current)
            current = {"marker": raw, "lines": []}
        elif current is not None:
            current["lines"].append(raw)
    if current is not None:
        steps.append(current)
    for idx, step in enumerate(steps, start=1):
        step["index"] = idx
    return steps


def format_number(value: float) -> str:
    if abs(value) < 5e-8:
        return "0"
    rounded = round(value)
    if abs(value - rounded) < 5e-6:
        return str(int(rounded))
    return f"{value:.8g}"


def stable_outputs_for_test(steps: list[dict], outputs: list[str]) -> set[str]:
    allowed = set(outputs)
    for step in steps:
        stable = step.get("stable", {})
        for output in outputs:
            if stable.get(output) is False:
                allowed.discard(output)
    return allowed


def build_action(
    action: str,
    step_results: dict[int, dict],
    outputs: list[str],
    allowed_outputs: set[str],
) -> str:
    rebuilt: list[str] = []
    steps = parse_steps(action)
    for position, step in enumerate(steps):
        rebuilt.append(step["marker"])
        kept_lines: list[str] = []
        for line in step["lines"]:
            if ANY_EXPECTED_RE.match(line):
                continue
            kept_lines.append(line)
            rebuilt.append(line)
        is_final_empty_delay = position == len(steps) - 1 and not any(line.strip() for line in kept_lines)
        if is_final_empty_delay:
            continue
        values = step_results.get(step["index"], {}).get("outputs", {})
        for output in outputs:
            if output in allowed_outputs and output in values:
                rebuilt.append(f"{output} = expValue({format_number(float(values[output]))});")
    return "\n".join(rebuilt)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--results", required=True)
    parser.add_argument("--outputs", required=True, help="Comma-separated root output names")
    parser.add_argument(
        "--exclude-outputs",
        default="",
        help="Comma-separated root outputs to remove from expValue backfill, for example unverified stateful outputs",
    )
    args = parser.parse_args()

    outputs = [name.strip() for name in args.outputs.split(",") if name.strip()]
    excluded_outputs = {name.strip() for name in args.exclude_outputs.split(",") if name.strip()}
    outputs = [name for name in outputs if name not in excluded_outputs]
    results = json.loads(Path(args.results).read_text(encoding="utf-8"))
    tests = results["tests"]
    if isinstance(tests, dict):
        tests = [tests]
    by_row = {}
    for item in tests:
        steps = item["steps"]
        if isinstance(steps, dict):
            steps = [steps]
        by_row[item["row"]] = {
            "steps": {step["index"]: step for step in steps},
            "allowed_outputs": stable_outputs_for_test(steps, outputs),
        }

    wb = load_workbook(args.workbook)
    ws = wb["TCSD"]
    for row, info in by_row.items():
        cell = ws.cell(row, 7)
        cell.value = build_action(cell.value or "", info["steps"], outputs, info["allowed_outputs"])
        alignment = copy(cell.alignment)
        alignment.wrap_text = True
        alignment.vertical = "top"
        cell.alignment = alignment
        line_count = cell.value.count("\n") + 1
        ws.row_dimensions[row].height = min(409, max(180, line_count * 13))
    wb.save(args.workbook)
    print(args.workbook)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
