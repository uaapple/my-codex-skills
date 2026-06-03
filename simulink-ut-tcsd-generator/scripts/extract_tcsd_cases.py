#!/usr/bin/env python3
"""Extract TCSD initializations/actions into JSON for simulation backfill."""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


STEP_RE = re.compile(r"^\s*\[\+\s*([0-9.]+)\s*(ms|s)\s*\](.*)$", re.IGNORECASE)
NUMBER = r"[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?"
VALUE = rf"({NUMBER}|\[\s*{NUMBER}(?:[\s,]+{NUMBER})*\s*\])"
PARAM_RE = re.compile(rf"^\s*p\s+([A-Za-z_]\w*)\s*=\s*{VALUE}\s*;")
ASSIGN_RE = re.compile(rf"^\s*([A-Za-z_]\w*)\s*=\s*{VALUE}\s*;")
INDEX_ASSIGN_RE = re.compile(rf"^\s*([A-Za-z_]\w*)\s+([1-9]\d*)\s*=\s*({NUMBER})\s*;")


def parse_value(text: str) -> float | list[float]:
    text = text.strip()
    if text.startswith("["):
        inner = text.strip("[]")
        return [float(item) for item in re.split(r"[\s,]+", inner.strip()) if item]
    return float(text)


def set_indexed_value(inputs: dict[str, object], name: str, index_text: str, value_text: str) -> None:
    index = int(index_text)
    value = float(value_text)
    current = inputs.get(name)
    values = list(current) if isinstance(current, list) else []
    while len(values) < index:
        values.append(0.0)
    values[index - 1] = value
    inputs[name] = values


def unknown_assignment(context: str, signal: str, line: str) -> dict[str, str]:
    return {"context": context, "signal": signal, "line": line}


def parse_assignments(
    text: str,
    input_names: set[str],
    context: str,
) -> tuple[dict[str, object], dict[str, object], list[dict[str, str]]]:
    inputs: dict[str, object] = {}
    params: dict[str, object] = {}
    unknowns: list[dict[str, str]] = []
    for raw in (text or "").splitlines():
        line = raw.strip()
        if not line or line.startswith("//"):
            continue
        param_match = PARAM_RE.match(line)
        if param_match:
            params[param_match.group(1)] = parse_value(param_match.group(2))
            continue
        index_match = INDEX_ASSIGN_RE.match(line)
        if index_match:
            if index_match.group(1) in input_names:
                set_indexed_value(inputs, index_match.group(1), index_match.group(2), index_match.group(3))
            else:
                unknowns.append(unknown_assignment(context, index_match.group(1), line))
            continue
        assign_match = ASSIGN_RE.match(line)
        if assign_match:
            if assign_match.group(1) in input_names:
                inputs[assign_match.group(1)] = parse_value(assign_match.group(2))
            else:
                unknowns.append(unknown_assignment(context, assign_match.group(1), line))
    return inputs, params, unknowns


def merge_assignments(
    base_inputs: dict[str, object],
    base_params: dict[str, object],
    override_inputs: dict[str, object],
    override_params: dict[str, object],
) -> tuple[dict[str, object], dict[str, object]]:
    """Merge TestGroup defaults with Test-specific assignments."""
    inputs = dict(base_inputs)
    params = dict(base_params)
    inputs.update(override_inputs)
    params.update(override_params)
    return inputs, params


def delay_seconds(amount: str, unit: str) -> float:
    value = float(amount)
    return value / 1000.0 if unit.lower() == "ms" else value


def parse_steps(
    action: str,
    input_names: set[str],
    row: int,
    test_id: str,
) -> tuple[list[dict], list[dict[str, str]]]:
    steps: list[dict] = []
    unknowns: list[dict[str, str]] = []
    current: dict | None = None
    for raw in (action or "").splitlines():
        match = STEP_RE.match(raw)
        if match:
            if current is not None:
                steps.append(current)
            current = {
                "marker": raw,
                "delay_s": delay_seconds(match.group(1), match.group(2)),
                "input_updates": {},
                "param_updates": {},
            }
            continue
        if current is None:
            continue
        context = f"row {row} test {test_id} action step {len(steps) + 1}"
        inputs, params, line_unknowns = parse_assignments(raw, input_names, context)
        unknowns.extend(line_unknowns)
        current["input_updates"].update(inputs)
        current["param_updates"].update(params)
    if current is not None:
        steps.append(current)
    for idx, step in enumerate(steps, start=1):
        step["index"] = idx
    return steps, unknowns


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--inputs", required=True, help="Comma-separated root input names")
    parser.add_argument("--output", required=True)
    parser.add_argument(
        "--allow-unknown-assignments",
        action="store_true",
        help="Ignore executable assignments whose left-hand side is not a root input",
    )
    args = parser.parse_args()

    input_names = {name.strip() for name in args.inputs.split(",") if name.strip()}
    wb = load_workbook(args.workbook)
    ws = wb["TCSD"]
    group_inputs: dict[str, object] = {}
    group_params: dict[str, object] = {}
    unknown_assignments: list[dict[str, str]] = []
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 3).value != "TestGroup":
            continue
        for offset in range(0, 2):
            context = f"row {row + offset} TestGroup initialization"
            inputs, params, unknowns = parse_assignments(ws.cell(row + offset, 6).value or "", input_names, context)
            unknown_assignments.extend(unknowns)
            group_inputs.update(inputs)
            group_params.update(params)
        break

    tests = []
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 3).value != "Test":
            continue
        test_id = ws.cell(row, 1).value
        context = f"row {row} test {test_id} initialization"
        test_inputs, test_params, unknowns = parse_assignments(ws.cell(row, 6).value or "", input_names, context)
        unknown_assignments.extend(unknowns)
        init_inputs, init_params = merge_assignments(group_inputs, group_params, test_inputs, test_params)
        steps, step_unknowns = parse_steps(ws.cell(row, 7).value or "", input_names, row, str(test_id or ""))
        unknown_assignments.extend(step_unknowns)
        tests.append(
            {
                "row": row,
                "test_id": test_id,
                "name": ws.cell(row, 2).value,
                "init_values": init_inputs,
                "init_params": init_params,
                "steps": steps,
            }
        )
    if unknown_assignments and not args.allow_unknown_assignments:
        print(
            json.dumps({"error": "unknown_tcsd_input_assignments", "items": unknown_assignments}, ensure_ascii=False, indent=2),
            file=sys.stderr,
        )
        return 1
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps({"tests": tests}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
