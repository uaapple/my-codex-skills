#!/usr/bin/env python3
"""Validate TCSD workbook signal names against compiled root ports."""

from __future__ import annotations

import argparse
import json
import re
import sys
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


STEP_RE = re.compile(r"^\s*\[\+\s*[0-9.]+\s*(ms|s)\s*\](?:\s*//.*)?$", re.IGNORECASE)
PARAM_ASSIGN_RE = re.compile(r"^\s*p\s+([A-Za-z_]\w*)\s*=")
EXP_ASSIGN_RE = re.compile(r"^\s*([A-Za-z_]\w*)\s*=\s*expValue\s*\(")
MACRO_EXP_ASSIGN_RE = re.compile(r"^\s*(\$output\[\d+\]\$)\s*=\s*expValue\s*\(")
INDEX_ASSIGN_RE = re.compile(r"^\s*([A-Za-z_]\w*)\s+([1-9]\d*)\s*=")
ASSIGN_RE = re.compile(r"^\s*([A-Za-z_]\w*)\s*=")
VECTOR_ASSIGN_RE = re.compile(r"^\s*([A-Za-z_]\w*)\s*=\s*\[")
FINAL_DELAY_RE = re.compile(r"^\s*\[\+\s*[0-9.]+\s*(ms|s)\s*\](?:\s*//.*)?\s*$", re.IGNORECASE)


def parse_csv_names(text: str | None) -> set[str]:
    if not text:
        return set()
    return {item.strip() for item in text.split(",") if item.strip()}


def normalize_name_items(items: Any) -> set[str]:
    names: set[str] = set()
    if not items:
        return names
    for item in items:
        if isinstance(item, dict):
            name = item.get("name")
        else:
            name = item
        if name:
            names.add(str(name).strip())
    return {name for name in names if name}


def load_interface_names(
    interface_json: str | None = None,
    inputs: str | None = None,
    outputs: str | None = None,
) -> tuple[set[str], set[str]]:
    root_inputs = parse_csv_names(inputs)
    root_outputs = parse_csv_names(outputs)
    if not interface_json:
        return root_inputs, root_outputs

    data = json.loads(Path(interface_json).read_text(encoding="utf-8"))
    if "rootPorts" in data:
        root_inputs.update(normalize_name_items(data["rootPorts"].get("inputs", [])))
        root_outputs.update(normalize_name_items(data["rootPorts"].get("outputs", [])))
    else:
        root_inputs.update(normalize_name_items(data.get("inputs", [])))
        root_outputs.update(normalize_name_items(data.get("outputs", [])))
    return root_inputs, root_outputs


def header_index(ws, header: str, fallback: int) -> int:
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(1, col).value or "").strip() == header:
            return col
    return fallback


def check_zip_integrity(workbook: Path, errors: list[dict[str, Any]]) -> None:
    if not zipfile.is_zipfile(workbook):
        errors.append({"code": "not_xlsx_zip", "workbook": str(workbook)})
        return
    try:
        with zipfile.ZipFile(workbook) as archive:
            bad_member = archive.testzip()
        if bad_member:
            errors.append({"code": "corrupt_xlsx_member", "workbook": str(workbook), "member": bad_member})
    except zipfile.BadZipFile as exc:
        errors.append({"code": "bad_xlsx_zip", "workbook": str(workbook), "message": str(exc)})


def error_record(
    code: str,
    sheet: str,
    cell: str,
    row: int,
    test_id: str,
    signal: str,
    line: str,
) -> dict[str, Any]:
    return {
        "code": code,
        "sheet": sheet,
        "cell": cell,
        "row": row,
        "test_id": test_id,
        "signal": signal,
        "line": line,
    }


def scan_cell(
    text: str,
    *,
    ws_title: str,
    cell: str,
    row: int,
    test_id: str,
    root_inputs: set[str],
    root_outputs: set[str],
) -> tuple[list[dict[str, Any]], int, int]:
    errors: list[dict[str, Any]] = []
    input_assignment_count = 0
    exp_count = 0

    for raw in (text or "").replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        line = raw.strip()
        if not line or line.startswith("//") or STEP_RE.match(line):
            continue
        if PARAM_ASSIGN_RE.match(line):
            continue

        macro_exp_match = MACRO_EXP_ASSIGN_RE.match(line)
        if macro_exp_match:
            exp_count += 1
            signal = macro_exp_match.group(1)
            errors.append(
                error_record("unsupported_vector_output_expectation", ws_title, cell, row, test_id, signal, line)
            )
            continue

        exp_match = EXP_ASSIGN_RE.match(line)
        if exp_match:
            exp_count += 1
            signal = exp_match.group(1)
            if signal not in root_outputs:
                errors.append(
                    error_record("unknown_exp_output", ws_title, cell, row, test_id, signal, line)
                )
            continue

        vector_match = VECTOR_ASSIGN_RE.match(line)
        if vector_match:
            signal = vector_match.group(1)
            if signal not in root_inputs:
                errors.append(
                    error_record("unknown_input_assignment", ws_title, cell, row, test_id, signal, line)
                )
            else:
                errors.append(
                    error_record("whole_vector_input_assignment", ws_title, cell, row, test_id, signal, line)
                )
            continue

        index_match = INDEX_ASSIGN_RE.match(line)
        if index_match:
            input_assignment_count += 1
            signal = index_match.group(1)
            if signal not in root_inputs:
                errors.append(
                    error_record("unknown_input_assignment", ws_title, cell, row, test_id, signal, line)
                )
            continue

        assign_match = ASSIGN_RE.match(line)
        if assign_match:
            input_assignment_count += 1
            signal = assign_match.group(1)
            if signal not in root_inputs:
                errors.append(
                    error_record("unknown_input_assignment", ws_title, cell, row, test_id, signal, line)
                )

    return errors, input_assignment_count, exp_count


def validate_workbook(
    workbook: str | Path,
    root_inputs: set[str],
    root_outputs: set[str],
    *,
    require_exp_values: bool = False,
) -> dict[str, Any]:
    workbook = Path(workbook)
    errors: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    check_zip_integrity(workbook, errors)
    if errors:
        return {"status": "failed", "errors": errors, "warnings": warnings}

    wb = load_workbook(workbook, data_only=False)
    if "TCSD" not in wb.sheetnames:
        return {
            "status": "failed",
            "errors": [{"code": "missing_tcsd_sheet", "workbook": str(workbook)}],
            "warnings": warnings,
        }

    ws = wb["TCSD"]
    test_id_col = header_index(ws, "TestID", 1)
    type_col = header_index(ws, "Type", 3)
    init_col = header_index(ws, "Initialization", 6)
    action_col = header_index(ws, "Action", 7)
    input_assignment_count = 0
    exp_count = 0
    test_count = 0

    for row in range(1, ws.max_row + 1):
        row_type = str(ws.cell(row, type_col).value or "").strip()
        if row_type not in {"Test", "TestGroup"}:
            continue
        if row_type == "Test":
            test_count += 1
        test_id = str(ws.cell(row, test_id_col).value or "").strip()
        for col in (init_col, action_col):
            cell_ref = ws.cell(row, col).coordinate
            cell_errors, input_seen, exp_seen = scan_cell(
                str(ws.cell(row, col).value or ""),
                ws_title=ws.title,
                cell=cell_ref,
                row=row,
                test_id=test_id,
                root_inputs=root_inputs,
                root_outputs=root_outputs,
            )
            errors.extend(cell_errors)
            input_assignment_count += input_seen
            exp_count += exp_seen

        if row_type == "Test":
            action_lines = [
                line.strip()
                for line in str(ws.cell(row, action_col).value or "").splitlines()
                if line.strip()
            ]
            if not action_lines or not FINAL_DELAY_RE.match(action_lines[-1]):
                errors.append(
                    {
                        "code": "missing_final_delay",
                        "sheet": ws.title,
                        "cell": ws.cell(row, action_col).coordinate,
                        "row": row,
                        "test_id": test_id,
                        "line": action_lines[-1] if action_lines else "",
                    }
                )

    if require_exp_values and exp_count == 0:
        errors.append({"code": "missing_exp_values", "workbook": str(workbook)})
    if not root_inputs:
        warnings.append({"code": "empty_root_input_set"})
    if not root_outputs:
        warnings.append({"code": "empty_root_output_set"})

    return {
        "status": "failed" if errors else "passed",
        "workbook": str(workbook),
        "test_count": test_count,
        "input_assignment_count": input_assignment_count,
        "exp_value_count": exp_count,
        "errors": errors,
        "warnings": warnings,
    }


def print_text_report(report: dict[str, Any]) -> None:
    if report["status"] == "passed":
        print(
            "TCSD workbook validation passed: "
            f"{report.get('test_count', 0)} tests, "
            f"{report.get('input_assignment_count', 0)} input assignments, "
            f"{report.get('exp_value_count', 0)} expValue expectations."
        )
        return

    print("TCSD workbook validation failed:", file=sys.stderr)
    for error in report.get("errors", []):
        location = ""
        if error.get("sheet") and error.get("cell"):
            location = f"{error['sheet']}!{error['cell']}"
        elif error.get("workbook"):
            location = error["workbook"]
        print(
            f"- {error.get('code')}: {location} "
            f"{error.get('test_id', '')} {error.get('signal', '')} {error.get('line', '')}".rstrip(),
            file=sys.stderr,
        )


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--inputs", help="Comma-separated compiled root Inport names")
    parser.add_argument("--outputs", help="Comma-separated root Outport names")
    parser.add_argument("--interface-json", help="JSON containing inputs/outputs or rootPorts.inputs/rootPorts.outputs")
    parser.add_argument("--require-exp-values", action="store_true")
    parser.add_argument("--report-json")
    args = parser.parse_args()

    root_inputs, root_outputs = load_interface_names(args.interface_json, args.inputs, args.outputs)
    report = validate_workbook(
        args.workbook,
        root_inputs,
        root_outputs,
        require_exp_values=args.require_exp_values,
    )
    if args.report_json:
        Path(args.report_json).write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    else:
        print_text_report(report)
    return 0 if report["status"] == "passed" else 1


if __name__ == "__main__":
    raise SystemExit(main())
