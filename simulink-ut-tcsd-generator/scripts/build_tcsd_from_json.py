#!/usr/bin/env python3
"""Build a TCSD workbook from a compact JSON spec and the bundled template."""

from __future__ import annotations

import argparse
import json
import re
from copy import copy
from pathlib import Path

from openpyxl import load_workbook

from validate_tcsd_workbook import load_interface_names, print_text_report, validate_workbook


HEADERS = [
    "TestID",
    "Name",
    "Type",
    "Requirement ID",
    "Test Case Description",
    "Initialization",
    "Action",
    "Work Status",
    "Report Links",
]

VECTOR_ASSIGN_RE = re.compile(r"^(\s*)([A-Za-z_]\w*)\s*=\s*\[([^\]]+)\]\s*;\s*$")
FINAL_DELAY_RE = re.compile(r"^\[\+\s*[0-9.]+\s*(ms|s)\s*\](?:\s*//.*)?$", re.IGNORECASE)
ASSIGNMENT_KEY_RE = re.compile(r"^\s*(p\s+)?([A-Za-z_]\w*)(?:\s+([1-9]\d*))?\s*=")


def expand_vector_assignments(text: str) -> str:
    """Convert whole-vector assignments to TCSD element assignments."""
    lines: list[str] = []
    for line in (text or "").splitlines():
        match = VECTOR_ASSIGN_RE.match(line)
        if not match:
            lines.append(line)
            continue
        indent, name, values_text = match.groups()
        values = [value for value in re.split(r"[\s,]+", values_text.strip()) if value]
        for index, value in enumerate(values, start=1):
            lines.append(f"{indent}{name} {index}={value};")
    return "\n".join(lines)


def ensure_final_delay(action: str, delay: str = "[+0.1s]") -> str:
    """Ensure each Test has a final run/sampling delay marker."""
    lines = (action or "").splitlines()
    while lines and not lines[-1].strip():
        lines.pop()
    if lines and FINAL_DELAY_RE.match(lines[-1].strip()):
        return "\n".join(lines)
    lines.append(delay)
    return "\n".join(lines)


def assignment_key(line: str) -> tuple[str, str, str] | None:
    """Return a stable key for scalar, indexed vector, and parameter assignments."""
    match = ASSIGNMENT_KEY_RE.match(line)
    if not match:
        return None
    prefix, name, index = match.groups()
    return ((prefix or "").strip(), name, index or "")


def merge_initializations(common: str, override: str) -> str:
    """Merge common TestGroup init into a Test, with Test-specific values winning."""
    merged: list[str] = []
    positions: dict[tuple[str, str, str], int] = {}

    for raw in (common or "").splitlines():
        line = raw.rstrip()
        if not line.strip():
            continue
        key = assignment_key(line)
        if key is not None:
            positions[key] = len(merged)
        merged.append(line)

    for raw in (override or "").splitlines():
        line = raw.rstrip()
        if not line.strip():
            continue
        key = assignment_key(line)
        if key is not None and key in positions:
            merged[positions[key]] = line
        else:
            if key is not None:
                positions[key] = len(merged)
            merged.append(line)

    return "\n".join(merged)


def copy_row_style(ws, src_row: int, dst_row: int) -> None:
    for col in range(1, 10):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--template", required=True)
    parser.add_argument("--spec", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--inputs", help="Comma-separated compiled root Inport names for workbook validation")
    parser.add_argument("--outputs", help="Comma-separated root Outport names for workbook validation")
    parser.add_argument("--interface-json", help="JSON containing inputs/outputs or rootPorts inputs/outputs")
    parser.add_argument("--skip-workbook-validation", action="store_true")
    args = parser.parse_args()

    spec = json.loads(Path(args.spec).read_text(encoding="utf-8"))
    wb = load_workbook(args.template)
    ws = wb["TCSD"]

    for col, header in enumerate(HEADERS, 1):
        ws.cell(1, col).value = header

    group = spec.get("test_group", {})
    ws.cell(2, 1).value = group.get("id", "INFO_TG_001")
    ws.cell(2, 2).value = group.get("name", spec.get("model_name", "Model"))
    ws.cell(2, 3).value = "TestGroup"
    ws.cell(2, 4).value = "TestGroup"
    ws.cell(2, 5).value = group.get("description", "")
    ws.cell(2, 6).value = group.get("initialization_1", "")
    ws.cell(3, 6).value = group.get("initialization_2", "")
    ws.cell(4, 6).value = group.get("output_reference_1", "")
    ws.cell(5, 6).value = group.get("output_reference_2", "")
    common_initialization = expand_vector_assignments(
        "\n".join(
            item
            for item in [group.get("initialization_1", ""), group.get("initialization_2", "")]
            if item
        )
    )

    for row in range(6, max(ws.max_row, 40) + 1):
        for col in range(1, 10):
            ws.cell(row, col).value = None

    for row, test in enumerate(spec["tests"], start=6):
        copy_row_style(ws, 6 + ((row - 6) % 5), row)
        ws.cell(row, 1).value = test["id"]
        ws.cell(row, 2).value = test["name"]
        ws.cell(row, 3).value = "Test"
        ws.cell(row, 4).value = test.get("requirement_id")
        ws.cell(row, 5).value = test.get("description", "")
        test_initialization = expand_vector_assignments(test.get("initialization", ""))
        ws.cell(row, 6).value = merge_initializations(common_initialization, test_initialization)
        ws.cell(row, 7).value = ensure_final_delay(expand_vector_assignments(test.get("action", "")))
        ws.cell(row, 8).value = test.get("work_status", "reviewed")
        ws.cell(row, 9).value = test.get("report_links")
        line_count = max(
            (ws.cell(row, 6).value or "").count("\n") + 1,
            (ws.cell(row, 7).value or "").count("\n") + 1,
        )
        ws.row_dimensions[row].height = min(409, max(180, line_count * 13))

    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    if not args.skip_workbook_validation and (args.inputs or args.outputs or args.interface_json):
        root_inputs, root_outputs = load_interface_names(args.interface_json, args.inputs, args.outputs)
        report = validate_workbook(output, root_inputs, root_outputs, require_exp_values=False)
        if report["status"] != "passed":
            print_text_report(report)
            return 1
    print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
